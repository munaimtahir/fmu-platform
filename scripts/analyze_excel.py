#!/usr/bin/env python3
"""
Quick script to analyze Excel file structure
Uses openpyxl directly (lighter dependency)
"""

import sys
import os

# Add backend to path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'backend'))

try:
    from openpyxl import load_workbook
except ImportError:
    print("Installing openpyxl...")
    os.system("pip3 install openpyxl --user 2>&1")
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("ERROR: Could not import openpyxl. Please install it manually:")
        print("  pip3 install openpyxl")
        sys.exit(1)

def analyze_excel(file_path):
    """Analyze Excel file structure"""
    print(f"Reading Excel file: {file_path}")
    
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        sheet = wb.active
        
        print(f"\nSheet name: {sheet.title}")
        print(f"Total rows: {sheet.max_row}")
        print(f"Total columns: {sheet.max_column}")
        
        # Get header row
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value if cell.value else "")
        
        print(f"\nColumns found ({len(headers)}):")
        for i, header in enumerate(headers, 1):
            print(f"  {i}. {header}")
        
        # Show first few data rows
        print(f"\nFirst 5 data rows:")
        for row_idx in range(2, min(7, sheet.max_row + 1)):
            row_data = []
            for cell in sheet[row_idx]:
                value = cell.value
                if value is None:
                    value = ""
                elif isinstance(value, (int, float)):
                    value = str(value)
                else:
                    value = str(value)[:50]  # Truncate long values
                row_data.append(value)
            print(f"\nRow {row_idx}:")
            for i, (header, value) in enumerate(zip(headers, row_data)):
                if value:  # Only show non-empty values
                    print(f"  {header}: {value}")
        
        wb.close()
        return headers
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()
        return None

if __name__ == "__main__":
    file_path = os.path.join(os.path.dirname(__file__), "..", "Year 1 formatted data.xlsx")
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        sys.exit(1)
    
    analyze_excel(file_path)
